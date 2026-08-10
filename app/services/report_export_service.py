from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.schemas.report import SalesReportResponse

REPORT_HEADERS = [
    "Tarih",
    "Sipariş Sayısı",
    "Satış Toplamı",
    "İndirim Toplamı",
    "KDV Toplamı",
]


def generate_sales_report_xlsx(
    report: SalesReportResponse,
) -> bytes:
    output = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Satış Raporu"

    worksheet.merge_cells("A1:E1")
    worksheet["A1"] = "Satış Raporu"
    worksheet["A1"].font = Font(
        bold=True,
        size=16,
        color="FFFFFF",
    )
    worksheet["A1"].fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )
    worksheet["A1"].alignment = Alignment(
        horizontal="center",
    )

    worksheet.merge_cells("A2:E2")
    worksheet["A2"] = (
        f"{report.date_from.isoformat()} - "
        f"{report.date_to.isoformat()}"
    )
    worksheet["A2"].alignment = Alignment(
        horizontal="center",
    )

    worksheet.append([])
    worksheet.append(REPORT_HEADERS)

    for cell in worksheet[4]:
        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="5B9BD5",
        )
        cell.alignment = Alignment(
            horizontal="center",
        )

    for row in report.daily_sales:
        worksheet.append(
            [
                row.date.isoformat(),
                row.order_count,
                row.sales_total,
                row.discount_total,
                row.vat_total,
            ]
        )

    worksheet.append([])
    worksheet.append(
        [
            "GENEL TOPLAM",
            report.total_orders,
            report.total_sales,
            report.total_discount,
            report.total_vat,
        ]
    )

    total_row = worksheet.max_row
    for cell in worksheet[total_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )

    for row in worksheet.iter_rows(
        min_row=5,
        min_col=3,
        max_col=5,
    ):
        for cell in row:
            cell.number_format = '#,##0.00" ₺"'

    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 18
    worksheet.column_dimensions["C"].width = 20
    worksheet.column_dimensions["D"].width = 20
    worksheet.column_dimensions["E"].width = 20
    worksheet.freeze_panes = "A5"

    workbook.save(output)
    return output.getvalue()


def generate_sales_report_pdf(
    report: SalesReportResponse,
) -> bytes:
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title="Sales Report",
    )

    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Sales Report", styles["Title"]),
        Paragraph(
            (
                f"Date range: {report.date_from.isoformat()} - "
                f"{report.date_to.isoformat()}"
            ),
            styles["Normal"],
        ),
        Spacer(1, 8 * mm),
    ]

    table_data = [
        [
            "Date",
            "Order Count",
            "Sales Total",
            "Discount Total",
            "VAT Total",
        ]
    ]

    for row in report.daily_sales:
        table_data.append(
            [
                row.date.isoformat(),
                str(row.order_count),
                f"{row.sales_total:.2f} TRY",
                f"{row.discount_total:.2f} TRY",
                f"{row.vat_total:.2f} TRY",
            ]
        )

    table_data.append(
        [
            "TOTAL",
            str(report.total_orders),
            f"{report.total_sales:.2f} TRY",
            f"{report.total_discount:.2f} TRY",
            f"{report.total_vat:.2f} TRY",
        ]
    )

    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[
            42 * mm,
            38 * mm,
            48 * mm,
            48 * mm,
            48 * mm,
        ],
    )
    table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#1F4E78"),
                ),
                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white,
                ),
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold",
                ),
                (
                    "BACKGROUND",
                    (0, -1),
                    (-1, -1),
                    colors.HexColor("#D9EAF7"),
                ),
                (
                    "FONTNAME",
                    (0, -1),
                    (-1, -1),
                    "Helvetica-Bold",
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey,
                ),
                (
                    "ALIGN",
                    (1, 1),
                    (-1, -1),
                    "RIGHT",
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE",
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, 0),
                    8,
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, 0),
                    8,
                ),
            ]
        )
    )

    elements.append(table)
    document.build(elements)

    return output.getvalue()